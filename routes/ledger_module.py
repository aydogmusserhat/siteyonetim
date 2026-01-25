# routes/ledger_module.py
# Admin + SuperAdmin: Daire bazında aylık borç/ödeme ekstre ekranı + Excel/PDF export

import io
from datetime import date, timedelta
from decimal import Decimal
from collections import defaultdict
from typing import Optional

from flask import request, render_template, redirect, url_for, flash, session, send_file
from sqlalchemy import and_, or_, func

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from models import db
from models.user_model import User
from models.apartment_model import Apartment
from models.bill_model import Bill
from models.payment_model import Payment
from models.site_model import Site


def register_ledger_routes(admin_bp, admin_required, _get_current_admin, _parse_date_flex):
    """
    admin_routes.py içinde 1 kez çağır.
    Route'lar admin_bp'ye eklenir.
    """
    # burada artık admin_routes import YOK!
    """
    admin_routes.py içinde 1 kez çağır.
    Route'lar admin_bp'ye eklenir.
    """

    # ✅ Circular import olmaması için local import

    # -------------------------
    # Helpers (admin fonksiyonlarına bağlı oldukları için burada)
    # -------------------------
    def _month_range_from_str(month_str: str):
        month_str = (month_str or "").strip()
        if not month_str:
            return None, None
        y, m = month_str.split("-")
        y = int(y)
        m = int(m)
        start = date(y, m, 1)
        if m == 12:
            end = date(y + 1, 1, 1)
        else:
            end = date(y, m + 1, 1)
        return start, end

    def _resolve_period_from_request():
        """
        Öncelik:
          1) month=YYYY-MM
          2) from=... to=... (esnek parse)
          3) yoksa bu ay
        """
        month = (request.args.get("month") or "").strip()
        if month:
            s, e = _month_range_from_str(month)
            return month, s, e

        date_from = (request.args.get("from") or "").strip()
        date_to = (request.args.get("to") or "").strip()

        if date_from and date_to:
            s = _parse_date_flex(date_from)
            e_inclusive = _parse_date_flex(date_to)
            e = e_inclusive + timedelta(days=1)  # inclusive -> exclusive
            return "", s, e

        # default this month
        today = date.today()
        s = date(today.year, today.month, 1)
        if today.month == 12:
            e = date(today.year + 1, 1, 1)
            month_label = f"{today.year}-12"
        else:
            e = date(today.year, today.month + 1, 1)
            month_label = f"{today.year}-{today.month:02d}"
        return month_label, s, e

    def _get_site_scope_for_admin(admin_user: User):
        """
        Admin => site zorunlu
        Super admin => aktif site seçiliyse o site, değilse global
        """
        site_id = session.get("active_site_id") or (admin_user.site_id if admin_user.site_id else None)
        is_super = (getattr(admin_user, "role", "") == "super_admin")
        global_mode = is_super and not site_id
        return site_id, global_mode

    def _ledger_build_dataset(
        site_id: Optional[int],
        global_mode: bool,
        start: date,
        end: date,
        apartment_id: Optional[int],
        q: Optional[str]
    ):
        """
        start-end aralığında:
          - billed: dönem borç toplamı
          - paid: dönem ödeme toplamı
          - open_balance: genel açık bakiye (open/partial)
          - bills/payments: dönem detay listeleri
        """

        # --- daire temel sorgusu ---
        apt_q = Apartment.query
        if not global_mode:
            apt_q = apt_q.filter(Apartment.site_id == site_id)

        if apartment_id:
            apt_q = apt_q.filter(Apartment.id == apartment_id)

        if q:
            like = f"%{q}%"
            # ⚠️ Bu alan adları sende farklıysa burayı kendi model alanlarına göre değiştir
            apt_q = apt_q.filter(
                (Apartment.block.ilike(like)) |
                (Apartment.floor.ilike(like)) |
                (Apartment.number.ilike(like)) |
                (Apartment.owner_name.ilike(like))
            )

        apartments = apt_q.order_by(
            Apartment.block.asc(),
            Apartment.floor.asc(),
            Apartment.number.asc()
        ).all()

        # --- her bill için toplam ödenen subquery ---
        pay_sum_subq = (
            db.session.query(
                Payment.bill_id.label("bill_id"),
                func.coalesce(func.sum(Payment.amount), 0).label("paid_sum"),
            )
            .group_by(Payment.bill_id)
            .subquery()
        )

        # --- billed (period) ---
        billed_q = db.session.query(
            Bill.apartment_id.label("apartment_id"),
            func.coalesce(func.sum(Bill.amount), 0).label("billed_sum")
        )

        if not global_mode:
            billed_q = billed_q.filter(Bill.site_id == site_id)

        billed_q = billed_q.filter(
            or_(
                and_(Bill.due_date.isnot(None), Bill.due_date >= start, Bill.due_date < end),
                and_(Bill.due_date.is_(None), Bill.created_at >= start, Bill.created_at < end),
            )
        ).group_by(Bill.apartment_id)

        billed_map = {r.apartment_id: r.billed_sum for r in billed_q.all()}

        # --- paid (period) ---
        paid_q = db.session.query(
            Payment.apartment_id.label("apartment_id"),
            func.coalesce(func.sum(Payment.amount), 0).label("paid_sum")
        )

        if not global_mode:
            paid_q = paid_q.filter(Payment.site_id == site_id)

        paid_q = paid_q.filter(
            Payment.payment_date >= start,
            Payment.payment_date < end,
        ).group_by(Payment.apartment_id)

        paid_map = {r.apartment_id: r.paid_sum for r in paid_q.all()}

        # --- open balance (overall) ---
        open_q = (
            db.session.query(
                Bill.apartment_id.label("apartment_id"),
                func.coalesce(
                    func.sum(Bill.amount - func.coalesce(pay_sum_subq.c.paid_sum, 0)),
                    0
                ).label("open_sum")
            )
            .outerjoin(pay_sum_subq, pay_sum_subq.c.bill_id == Bill.id)
            .filter(Bill.status.in_(["open", "partial"]))
        )

        if not global_mode:
            open_q = open_q.filter(Bill.site_id == site_id)

        open_q = open_q.group_by(Bill.apartment_id)
        open_map = {r.apartment_id: r.open_sum for r in open_q.all()}

        # --- detaylar (period içinde borç/ödeme listesi) ---
        bill_detail_q = db.session.query(Bill).join(Apartment, Bill.apartment_id == Apartment.id)
        if not global_mode:
            bill_detail_q = bill_detail_q.filter(Bill.site_id == site_id)
        bill_detail_q = bill_detail_q.filter(
            or_(
                and_(Bill.due_date.isnot(None), Bill.due_date >= start, Bill.due_date < end),
                and_(Bill.due_date.is_(None), Bill.created_at >= start, Bill.created_at < end),
            )
        )
        if apartment_id:
            bill_detail_q = bill_detail_q.filter(Bill.apartment_id == apartment_id)

        bills = bill_detail_q.order_by(Bill.due_date.desc().nullslast(), Bill.created_at.desc()).all()
        bills_by_apt = defaultdict(list)
        for b in bills:
            bills_by_apt[b.apartment_id].append(b)

        pay_detail_q = db.session.query(Payment).join(Apartment, Payment.apartment_id == Apartment.id)
        if not global_mode:
            pay_detail_q = pay_detail_q.filter(Payment.site_id == site_id)
        pay_detail_q = pay_detail_q.filter(Payment.payment_date >= start, Payment.payment_date < end)
        if apartment_id:
            pay_detail_q = pay_detail_q.filter(Payment.apartment_id == apartment_id)

        payments = pay_detail_q.order_by(Payment.payment_date.desc()).all()
        pays_by_apt = defaultdict(list)
        for p in payments:
            pays_by_apt[p.apartment_id].append(p)

        # --- output rows ---
        rows = []
        for apt in apartments:
            billed = Decimal(str(billed_map.get(apt.id, 0) or 0))
            paid = Decimal(str(paid_map.get(apt.id, 0) or 0))
            open_bal = Decimal(str(open_map.get(apt.id, 0) or 0))
            rows.append({
                "apartment": apt,
                "site_id": getattr(apt, "site_id", None),
                "billed": billed,
                "paid": paid,
                "net": paid - billed,
                "open_balance": open_bal,
                "bills": bills_by_apt.get(apt.id, []),
                "payments": pays_by_apt.get(apt.id, []),
            })

        site_map = {}
        if global_mode:
            sites = Site.query.order_by(Site.name.asc()).all()
            site_map = {s.id: s.name for s in sites}

        return rows, site_map

    # -------------------------
    # Routes
    # -------------------------
    @admin_bp.route("/ledger", methods=["GET"])
    @admin_required
    def apartment_ledger():
        admin_user = _get_current_admin()
        if not admin_user:
            flash("Kullanıcı bulunamadı. Lütfen tekrar giriş yapın.", "error")
            return redirect(url_for("auth.logout"))

        site_id, global_mode = _get_site_scope_for_admin(admin_user)
        if not global_mode and not site_id:
            flash("Bu ekran için bir siteye atanmış olmanız gerekiyor.", "error")
            return redirect(url_for("admin.dashboard"))

        month_label, start, end = _resolve_period_from_request()
        apartment_id = request.args.get("apartment_id", type=int)
        q = (request.args.get("q") or "").strip() or None

        rows, site_map = _ledger_build_dataset(site_id, global_mode, start, end, apartment_id, q)

        apt_list_q = Apartment.query
        if not global_mode:
            apt_list_q = apt_list_q.filter_by(site_id=site_id)
        apt_list = apt_list_q.order_by(Apartment.block.asc(), Apartment.floor.asc(), Apartment.number.asc()).all()
        args = request.args.to_dict(flat=True)
        period_end_inclusive = end - timedelta(days=1)
        return render_template(
            "admin/ledger.html",
            args=args,
            period_end_inclusive=period_end_inclusive,
            admin_user=admin_user,
            global_mode=global_mode,
            month_label=month_label,
            period_start=start,
            period_end=end,
            rows=rows,
            site_map=site_map,
            apartments=apt_list,
            selected_apartment_id=apartment_id,
            q=q or "",
        )

    @admin_bp.route("/ledger/export.xlsx", methods=["GET"])
    @admin_required
    def apartment_ledger_export_xlsx():
        admin_user = _get_current_admin()
        site_id, global_mode = _get_site_scope_for_admin(admin_user)

        if not global_mode and not site_id:
            flash("Bu işlem için bir siteye atanmış olmanız gerekiyor.", "error")
            return redirect(url_for("admin.dashboard"))

        month_label, start, end = _resolve_period_from_request()
        apartment_id = request.args.get("apartment_id", type=int)
        q = (request.args.get("q") or "").strip() or None

        rows, site_map = _ledger_build_dataset(site_id, global_mode, start, end, apartment_id, q)

        wb = Workbook()
        ws = wb.active
        ws.title = "Ekstre"

        headers = []
        if global_mode:
            headers.append("Site")
        headers += ["Daire", "Malik", "Dönem Borç", "Dönem Ödeme", "Net (Ödeme-Borç)", "Açık Bakiye (Genel)"]
        ws.append(headers)

        for r in rows:
            apt = r["apartment"]
            apt_label = f"{getattr(apt,'block','')}-{getattr(apt,'floor','')}-{getattr(apt,'number','')}".strip("-")
            owner = getattr(apt, "owner_name", "") or ""

            line = []
            if global_mode:
                line.append(site_map.get(r["site_id"], ""))
            line += [
                apt_label,
                owner,
                float(r["billed"]),
                float(r["paid"]),
                float(r["net"]),
                float(r["open_balance"]),
            ]
            ws.append(line)

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"rapor_{month_label or start.isoformat()}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @admin_bp.route("/ledger/export.pdf", methods=["GET"])
    @admin_required
    def apartment_ledger_export_pdf():
        admin_user = _get_current_admin()
        site_id, global_mode = _get_site_scope_for_admin(admin_user)

        if not global_mode and not site_id:
            flash("Bu işlem için bir siteye atanmış olmanız gerekiyor.", "error")
            return redirect(url_for("admin.dashboard"))

        month_label, start, end = _resolve_period_from_request()
        apartment_id = request.args.get("apartment_id", type=int)
        q = (request.args.get("q") or "").strip() or None

        rows, site_map = _ledger_build_dataset(site_id, global_mode, start, end, apartment_id, q)

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        title = f"Rapor ({start.isoformat()} - {(end - timedelta(days=1)).isoformat()})"
        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, height - 50, title)

        y = height - 80
        c.setFont("Helvetica", 9)

        if global_mode:
            cols = [("Site", 40), ("Daire", 140), ("Borç", 260), ("Ödeme", 330), ("Net", 400), ("Açık", 470)]
        else:
            cols = [("Daire", 40), ("Borç", 200), ("Ödeme", 280), ("Net", 360), ("Açık", 440)]

        for name, x in cols:
            c.drawString(x, y, name)

        y -= 14
        c.line(40, y, width - 40, y)
        y -= 10

        for r in rows:
            if y < 60:
                c.showPage()
                y = height - 60
                c.setFont("Helvetica", 9)

            apt = r["apartment"]
            apt_label = f"{getattr(apt,'block','')}-{getattr(apt,'floor','')}-{getattr(apt,'number','')}".strip("-")

            if global_mode:
                c.drawString(40, y, (site_map.get(r["site_id"], "") or "")[:18])
                c.drawString(140, y, apt_label[:18])
                c.drawRightString(305, y, f"{r['billed']:.2f}")
                c.drawRightString(375, y, f"{r['paid']:.2f}")
                c.drawRightString(445, y, f"{r['net']:.2f}")
                c.drawRightString(width - 40, y, f"{r['open_balance']:.2f}")
            else:
                c.drawString(40, y, apt_label[:18])
                c.drawRightString(255, y, f"{r['billed']:.2f}")
                c.drawRightString(335, y, f"{r['paid']:.2f}")
                c.drawRightString(415, y, f"{r['net']:.2f}")
                c.drawRightString(width - 40, y, f"{r['open_balance']:.2f}")

            y -= 12

        c.save()
        buffer.seek(0)

        filename = f"rapor_{month_label or start.isoformat()}.pdf"
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")
